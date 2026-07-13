import logging
import os
import sqlite3
from datetime import datetime

import pandas as pd

from classify.taxonomy import ISIC_REV5

log = logging.getLogger("classify.reporter")


def _segment_stats(projects_df: pd.DataFrame) -> pd.DataFrame:
    """Project counts per (type, repository) segment, broken down by status."""
    if projects_df.empty:
        return pd.DataFrame()
    df = projects_df.copy()
    df["type"] = df["type"].fillna("UNKNOWN")
    df["classification_status"] = df["classification_status"].fillna("UNCLASSIFIED")
    seg = (
        df.groupby(["type", "repository", "classification_status"])
        .size()
        .unstack(fill_value=0)
        .reset_index()
    )
    status_cols = [c for c in seg.columns if c not in ("type", "repository")]
    seg["total"] = seg[status_cols].sum(axis=1)
    return seg.sort_values(["type", "repository"])


class ClassificationReporter:
    def __init__(self, db_path: str, output_dir: str):
        self.db_path = db_path
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    def generate_all(self, run_timestamp: str | None = None) -> dict[str, str]:
        ts = run_timestamp or datetime.now().strftime("%Y%m%d_%H%M%S")
        paths = {}
        try:
            paths["excel"] = self.generate_excel(ts)
        except Exception as exc:
            log.error("Excel generation failed: %s", exc, exc_info=True)
        try:
            paths["pdf"] = self.generate_pdf(ts)
        except Exception as exc:
            log.error("PDF generation failed: %s", exc, exc_info=True)
        try:
            paths["review_csv"] = self.generate_review_csv(ts)
        except Exception as exc:
            log.error("Review CSV generation failed: %s", exc, exc_info=True)
        try:
            file_csv = self.generate_file_review_csv(ts)
            if file_csv:
                paths["file_review_csv"] = file_csv
        except Exception as exc:
            log.error("File review CSV generation failed: %s", exc, exc_info=True)
        return paths

    # ── Excel ─────────────────────────────────────────────────────────────────

    def generate_excel(self, timestamp: str) -> str:
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows

        path = os.path.join(self.output_dir, f"qda_classification_{timestamp}.xlsx")
        wb = Workbook()
        wb.remove(wb.active)

        with sqlite3.connect(self.db_path) as conn:
            projects_df = pd.read_sql_query(
                """
                SELECT p.id, p.title, p.type, p.class,
                       c.repository,
                       c.isic_section, d.section_title AS isic_section_title,
                       c.isic_division, d.division_title AS isic_division_title,
                       c.section_confidence, c.division_confidence,
                       c.classification_status,
                       c.secondary_class, c.secondary_section, c.secondary_division,
                       c.method, c.rerank_confidence,
                       c.summary, c.flags,
                       c.classified_at, p.project_url, p.doi
                FROM projects p
                LEFT JOIN classifications c
                       ON c.project_id = p.id AND c.entity_type = 'PROJECT'
                LEFT JOIN dim_isic d ON d.class = c.class
                ORDER BY p.id
                """,
                conn,
            )

            files_df = pd.read_sql_query(
                """
                SELECT f.id, f.project_id, p.title AS project_title,
                       p.type AS project_type,
                       p.download_repository_folder AS repository,
                       f.file_name, f.file_type, f.class, f.status
                FROM files f
                JOIN projects p ON f.project_id = p.id
                ORDER BY f.project_id, f.file_name
                """,
                conn,
            )

            file_class_df = self._file_classifications_df(conn)

            classified = projects_df[projects_df["isic_section"].notna()].copy()

        def write_sheet(name: str, df: pd.DataFrame) -> None:
            ws = wb.create_sheet(name[:31])
            for row in dataframe_to_rows(df, index=False, header=True):
                ws.append(row)

        write_sheet("Projects", projects_df)
        write_sheet("Files", files_df)

        if not classified.empty:
            sec_stats = (
                classified.groupby("isic_section")
                .agg(count=("id", "count"), section_title=("isic_section_title", "first"))
                .reset_index()
                .sort_values("count", ascending=False)
            )
            write_sheet("Stats_by_Section", sec_stats)

            div_stats = (
                classified.groupby(["isic_section", "isic_division"])
                .agg(
                    count=("id", "count"),
                    section_title=("isic_section_title", "first"),
                    division_title=("isic_division_title", "first"),
                )
                .reset_index()
                .sort_values("count", ascending=False)
            )
            write_sheet("Stats_by_Division", div_stats)
        else:
            write_sheet("Stats_by_Section", pd.DataFrame())
            write_sheet("Stats_by_Division", pd.DataFrame())

        write_sheet("Stats_by_Type_Repo", _segment_stats(projects_df))

        if not file_class_df.empty:
            write_sheet("File_Classifications", file_class_df)
            file_sec_stats = (
                file_class_df[file_class_df["isic_section"].notna()]
                .groupby(["project_type", "repository", "isic_section"])
                .agg(count=("file_id", "count"), section_title=("isic_section_title", "first"))
                .reset_index()
                .sort_values("count", ascending=False)
            )
            write_sheet("File_Stats_by_Section", file_sec_stats)

        flagged = projects_df[
            projects_df["classification_status"].isin(["NEEDS_REVIEW", "AMBIGUOUS"])
            | projects_df["flags"].str.contains("flagged_qda", na=False)
            | projects_df["classification_status"].isin(["INSUFFICIENT_DATA"])
        ].copy()
        write_sheet("Flagged_Review", flagged)

        # Histogram tab
        try:
            hist_path = self._build_histogram(classified, timestamp)
            if hist_path:
                from openpyxl.drawing.image import Image as XLImage
                ws_hist = wb.create_sheet("Histogram")
                ws_hist["A1"] = "Classification distribution"
                img = XLImage(hist_path)
                img.width, img.height = 900, 640
                ws_hist.add_image(img, "A3")
        except Exception as exc:
            log.warning("Histogram embedding failed: %s", exc)

        wb.save(path)
        log.info("Excel report saved: %s", path)
        return path

    def _file_classifications_df(self, conn) -> pd.DataFrame:
        try:
            return pd.read_sql_query(
                """
                SELECT c.file_id, c.project_id, p.title AS project_title,
                       p.type AS project_type, c.repository,
                       f.file_name, f.file_type, c.class,
                       c.isic_section, d.section_title AS isic_section_title,
                       c.isic_division, d.division_title AS isic_division_title,
                       c.section_confidence, c.division_confidence,
                       c.classification_status, c.flags
                FROM classifications c
                JOIN files f ON f.id = c.file_id
                JOIN projects p ON p.id = c.project_id
                LEFT JOIN dim_isic d ON d.class = c.class
                WHERE c.entity_type = 'FILE'
                ORDER BY c.project_id, f.file_name
                """,
                conn,
            )
        except Exception:
            return pd.DataFrame()  # not yet populated

    # ── PDF ───────────────────────────────────────────────────────────────────

    def generate_pdf(self, timestamp: str) -> str:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import (
            Image, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
        )

        path = os.path.join(self.output_dir, f"qda_classification_{timestamp}.pdf")

        with sqlite3.connect(self.db_path) as conn:
            total = pd.read_sql_query("SELECT COUNT(*) AS n FROM projects", conn).iloc[0]["n"]
            classified_n = pd.read_sql_query(
                "SELECT COUNT(*) AS n FROM classifications "
                "WHERE entity_type='PROJECT' AND isic_section IS NOT NULL", conn
            ).iloc[0]["n"]
            status_df = pd.read_sql_query(
                "SELECT classification_status, COUNT(*) AS n FROM classifications "
                "WHERE entity_type='PROJECT' GROUP BY classification_status",
                conn,
            )
            sec_df = pd.read_sql_query(
                """SELECT c.isic_section, d.section_title AS isic_section_title, COUNT(*) AS n
                   FROM classifications c
                   LEFT JOIN dim_isic d ON d.class = c.class
                   WHERE c.entity_type='PROJECT' AND c.isic_section IS NOT NULL
                   GROUP BY c.isic_section ORDER BY n DESC""",
                conn,
            )
            seg_df = pd.read_sql_query(
                """SELECT COALESCE(p.type, 'UNKNOWN') AS type,
                          p.download_repository_folder AS repository,
                          COALESCE(c.classification_status, 'UNCLASSIFIED') AS status,
                          COUNT(*) AS n
                   FROM projects p
                   LEFT JOIN classifications c
                          ON c.project_id = p.id AND c.entity_type = 'PROJECT'
                   GROUP BY 1, 2, 3
                   ORDER BY 1, 2""",
                conn,
            )

        styles = getSampleStyleSheet()
        H = ParagraphStyle("H", parent=styles["Heading1"], fontSize=16, spaceAfter=12)
        H2 = ParagraphStyle("H2", parent=styles["Heading2"], fontSize=12, spaceAfter=8)

        doc = SimpleDocTemplate(
            path, pagesize=A4,
            rightMargin=2 * cm, leftMargin=2 * cm,
            topMargin=2 * cm, bottomMargin=2 * cm,
        )

        story = [
            Paragraph("QDA Classification Report", H),
            Paragraph(
                f"Generated: {datetime.now():%Y-%m-%d %H:%M} | "
                f"Total projects: {total} | Classified: {classified_n}",
                styles["Normal"],
            ),
            Spacer(1, 0.5 * cm),
            Paragraph("Summary", H2),
        ]

        summary_data = [["Metric", "Value"]] + [
            [row["classification_status"], str(row["n"])]
            for _, row in status_df.iterrows()
        ]
        t = Table(summary_data, colWidths=[8 * cm, 5 * cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.5 * cm))

        if not sec_df.empty:
            story.append(Paragraph("ISIC Section Distribution", H2))
            sec_data = [["Section", "Title", "Count"]] + [
                [row["isic_section"], (row["isic_section_title"] or "")[:55], str(row["n"])]
                for _, row in sec_df.iterrows()
            ]
            t2 = Table(sec_data, colWidths=[2 * cm, 11 * cm, 2 * cm])
            t2.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgreen),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
            ]))
            story.append(t2)

        if not seg_df.empty:
            story.append(Spacer(1, 0.5 * cm))
            story.append(Paragraph("Projects by Type and Repository", H2))
            seg_data = [["Type", "Repository", "Status", "Count"]] + [
                [row["type"], row["repository"], row["status"], str(row["n"])]
                for _, row in seg_df.iterrows()
            ]
            t3 = Table(seg_data, colWidths=[4.5 * cm, 3.5 * cm, 5 * cm, 2 * cm])
            t3.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightblue),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
            ]))
            story.append(t3)

        # Histogram page
        with sqlite3.connect(self.db_path) as conn:
            cl_df = pd.read_sql_query(
                "SELECT isic_section, isic_division FROM classifications "
                "WHERE entity_type='PROJECT' AND isic_section IS NOT NULL",
                conn,
            )
        hist_path = self._build_histogram(cl_df, timestamp)
        if hist_path:
            story += [
                PageBreak(),
                Paragraph("Distribution Histograms", H2),
                Image(hist_path, width=17 * cm, height=12 * cm),
            ]

        doc.build(story)
        log.info("PDF report saved: %s", path)
        return path

    # ── Review CSV ────────────────────────────────────────────────────────────

    def generate_review_csv(self, timestamp: str) -> str:
        path = os.path.join(self.output_dir, f"review_queue_{timestamp}.csv")
        with sqlite3.connect(self.db_path) as conn:
            df = pd.read_sql_query(
                """
                SELECT p.id, p.title, p.type, c.repository, p.project_url,
                       c.class, c.isic_section, c.isic_division,
                       c.secondary_class, c.secondary_section, c.secondary_division,
                       c.section_confidence, c.division_confidence,
                       c.method, c.rerank_confidence,
                       c.classification_status, c.summary, c.flags
                FROM projects p
                LEFT JOIN classifications c
                       ON c.project_id = p.id AND c.entity_type = 'PROJECT'
                WHERE c.classification_status IN ('NEEDS_REVIEW','AMBIGUOUS','INSUFFICIENT_DATA')
                   OR c.flags LIKE '%flagged_qda%'
                   OR (c.id IS NULL)
                ORDER BY p.id
                """,
                conn,
            )
        df.to_csv(path, index=False)
        log.info("Review queue saved: %s (%d rows)", path, len(df))
        return path

    def generate_file_review_csv(self, timestamp: str) -> str | None:
        """Per-file review queue; None when no file classifications exist."""
        with sqlite3.connect(self.db_path) as conn:
            df = self._file_classifications_df(conn)
        if df.empty:
            return None
        flagged = df[df["classification_status"].isin(["NEEDS_REVIEW", "AMBIGUOUS", "ERROR"])]
        path = os.path.join(self.output_dir, f"file_review_queue_{timestamp}.csv")
        flagged.to_csv(path, index=False)
        log.info("File review queue saved: %s (%d rows)", path, len(flagged))
        return path

    # ── Histogram helper ──────────────────────────────────────────────────────

    def _build_histogram(self, df: pd.DataFrame, timestamp: str) -> str | None:
        try:
            import matplotlib
            matplotlib.use("Agg")
            import matplotlib.pyplot as plt

            fig, axes = plt.subplots(1, 2, figsize=(14, 5))

            ax = axes[0]
            if not df.empty and "isic_section" in df.columns:
                counts = df["isic_section"].value_counts().sort_index()
                counts.plot(kind="bar", ax=ax, color="steelblue")
            ax.set_title("Projects per ISIC Section")
            ax.set_xlabel("Section")
            ax.set_ylabel("Count")
            ax.tick_params(axis="x", rotation=0)

            ax = axes[1]
            if not df.empty and "isic_section" in df.columns and "isic_division" in df.columns:
                df = df.copy()
                df["sec_div"] = df["isic_section"].fillna("?") + "/" + df["isic_division"].fillna("?")
                top15 = df["sec_div"].value_counts().head(15)
                top15.plot(kind="barh", ax=ax, color="indianred")
            ax.set_title("Top ISIC Divisions")
            ax.set_xlabel("Count")

            plt.tight_layout()
            hist_path = os.path.join(self.output_dir, f"histogram_{timestamp}.png")
            plt.savefig(hist_path, dpi=150, bbox_inches="tight")
            plt.close(fig)
            return hist_path
        except Exception as exc:
            log.warning("Histogram creation failed: %s", exc)
            return None
